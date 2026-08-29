#!/usr/bin/env python3
"""Build tpt7_node_sample.xml: a FundsXML 4.x file whose RegulatoryReportings/
IndirectReporting/TripartiteTemplateSolvencyII_V7 (TPT7) node is populated from
a real example TPT file (example_files_tpt.zip).

This demonstrates the *native* FundsXML representation of a TPT report: unlike
the FundDynamicData look-through (which tpt_v7_export.xslt flattens), the TPT7
node already carries every TPT column as a structured element -- so the reverse
transform tpt7_node_to_csv.xslt is a near 1:1 field copy.

Source mapping (TPT column -> TPT7 element) follows the spec spreadsheet's
"Fundxml data name and path" column. We populate the mandatory elements plus the
issuer/QRT blocks and (for completeness) per-position credit-risk issuer data.

Usage:  python build_tpt7_sample.py <example.xlsx> <out.xml>
"""
import sys, os, warnings
warnings.filterwarnings("ignore")
import openpyxl
from lxml import etree

E = etree.SubElement


def num(v):
    """Render a spreadsheet value as a plain decimal/int string ('' if blank)."""
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def codif(v):
    """Codification system code as integer string (cells come back as 1.0/99.0)."""
    return num(v)


def build(xlsx, out):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    data = [r for r in rows[1:] if r[0]]
    col = {h.split("_")[0]: i for i, h in enumerate(hdr)}

    def c(row, n):
        return row[col[n]] if n in col else None

    p0 = data[0]  # portfolio-level values repeat on every row

    root = etree.Element("FundsXML4")
    root.set("{http://www.w3.org/2001/XMLSchema-instance}noNamespaceSchemaLocation",
             "FundsXML4.xsd")
    # nsmap fix: set xsi prefix
    etree.register_namespace  # noop, keep import tidy

    # --- ControlData ---
    cd = E(root, "ControlData")
    E(cd, "UniqueDocumentID").text = "TPT7_SAMPLE_" + str(c(p0, "1"))
    E(cd, "DocumentGenerated").text = str(c(p0, "7")) + "T00:00:00"
    E(cd, "ContentDate").text = str(c(p0, "7"))
    ds = E(cd, "DataSupplier")
    E(ds, "SystemCountry").text = "AT"
    E(ds, "Short").text = "EAM"
    E(ds, "Name").text = str(c(p0, "117"))
    E(ds, "Type").text = "IC"
    E(cd, "DataOperation").text = "INITIAL"
    E(cd, "Language").text = "EN"

    # --- minimal Fund (identification only) ---
    funds = E(root, "Funds")
    fund = E(funds, "Fund")
    ids = E(fund, "Identifiers")
    E(ids, "ISIN").text = str(c(p0, "1"))
    nm = E(fund, "Names")
    E(nm, "OfficialName").text = str(c(p0, "3"))
    E(fund, "Currency").text = str(c(p0, "4"))
    E(fund, "SingleFundFlag").text = "true"

    # --- RegulatoryReportings / IndirectReporting / TPT7 ---
    rr = E(root, "RegulatoryReportings")
    ind = E(rr, "IndirectReporting")
    tpt = E(ind, "TripartiteTemplateSolvencyII_V7")
    pf = E(tpt, "Portfolio")

    E(pf, "TPTVersion").text = "V7.0"
    pid = E(pf, "PortfolioID")
    E(pid, "CodificationSystem").text = codif(c(p0, "2"))
    E(pid, "Code").text = str(c(p0, "1"))
    E(pf, "PortfolioName").text = str(c(p0, "3"))
    E(pf, "PortfolioCurrency").text = str(c(p0, "4"))
    E(pf, "TotalNetAssets").text = num(c(p0, "5"))
    E(pf, "ValuationDate").text = str(c(p0, "6"))
    E(pf, "ReportingDate").text = str(c(p0, "7"))
    if num(c(p0, "8")):
        scn = E(pf, "ShareClass")
        E(scn, "SharePrice").text = num(c(p0, "8"))
        if num(c(p0, "8b")):
            E(scn, "TotalNumberOfShares").text = num(c(p0, "8b"))
    if num(c(p0, "9")):
        E(pf, "CashPercentage").text = num(c(p0, "9"))
    E(pf, "CompleteSCRDelivery").text = str(c(p0, "11") or "N")

    # QRTPortfolioInformation (fund issuer + group + CIC + custodian)
    qrt = E(pf, "QRTPortfolioInformation")
    fi = E(qrt, "FundIssuer")
    E(fi, "Name").text = str(c(p0, "117"))
    if num(c(p0, "115")):
        E(fi, "Code").text = str(c(p0, "115"))
    E(fi, "CodeType").text = codif(c(p0, "116")) or "1"
    E(fi, "EconomicSector").text = str(c(p0, "118"))
    E(fi, "Country").text = str(c(p0, "122"))
    fg = E(qrt, "FundIssuerGroup")
    E(fg, "Name").text = str(c(p0, "121"))
    if num(c(p0, "119")):
        E(fg, "Code").text = str(c(p0, "119"))
    E(fg, "CodeType").text = codif(c(p0, "120")) or "1"
    E(qrt, "FundCIC").text = str(c(p0, "123"))
    E(qrt, "FundCustodianCountry").text = str(c(p0, "123a"))

    # Positions
    poss = E(pf, "Positions")
    for r in data:
        pos = E(poss, "Position")
        E(pos, "InstrumentCIC").text = str(c(r, "12"))
        ic = E(pos, "InstrumentCode")
        E(ic, "CodificationSystem").text = codif(c(r, "15"))
        E(ic, "Code").text = str(c(r, "14"))
        E(pos, "InstrumentName").text = str(c(r, "17"))
        val = E(pos, "Valuation")
        if num(c(r, "18")):
            E(val, "Quantity").text = num(c(r, "18"))
        if num(c(r, "19")):
            E(val, "TotalNominalValueQC").text = num(c(r, "19"))
        E(val, "QuotationCurrency").text = str(c(r, "21"))
        E(val, "MarketValueQC").text = num(c(r, "22"))
        E(val, "CleanValueQC").text = num(c(r, "23"))
        E(val, "MarketValuePC").text = num(c(r, "24"))
        E(val, "CleanValuePC").text = num(c(r, "25"))
        E(val, "PositionWeight").text = num(c(r, "26"))
        E(val, "MarketExposureQC").text = num(c(r, "27"))
        E(val, "MarketExposurePC").text = num(c(r, "28"))
        E(val, "MarketExposureWeight").text = num(c(r, "30"))
        # Per-position credit-risk issuer data (cols 46/47/48, group, 52)
        if str(c(r, "46") or "").strip():
            crd = E(pos, "CreditRiskData")
            ii = E(crd, "InstrumentIssuer")
            E(ii, "Name").text = str(c(r, "46"))
            if num(c(r, "47")):
                E(ii, "Code").text = str(c(r, "47"))
            E(ii, "CodeType").text = codif(c(r, "48")) or "1"
            ig = E(crd, "IssuerGroup")
            E(ig, "Name").text = str(c(p0, "121"))
            if num(c(p0, "119")):
                E(ig, "Code").text = str(c(p0, "119"))
            E(ig, "CodeType").text = "1"
            E(crd, "IssuerCountry").text = str(c(r, "52"))
        # QRT position info: underlying asset category (col 131)
        if num(c(r, "131")):
            qpi = E(pos, "QRTPositionInformation")
            E(qpi, "UnderlyingAssetCategory").text = num(c(r, "131"))
        # AdditionalInformation is mandatory (all children optional -> empty)
        E(pos, "AdditionalInformation")

    tree = etree.ElementTree(root)
    tree.write(out, xml_declaration=True, encoding="UTF-8", pretty_print=True)
    print(f"wrote {out}: 1 portfolio, {len(data)} positions "
          f"(source {os.path.basename(xlsx)})")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("usage: build_tpt7_sample.py <example.xlsx> <out.xml>", file=sys.stderr)
        sys.exit(2)
    build(sys.argv[1], sys.argv[2])
